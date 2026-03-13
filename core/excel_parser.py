import re
from pathlib import Path

import openpyxl
import yaml

from core.models import MarketData, SegmentData, ExtendedMarketData, RegionData, ChartItem


def _safe_float(value, default: float = 0.0) -> float:
    """Safely convert a cell value to float, stripping %, whitespace, and stray characters."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("%", "").replace(",", "")
    # Remove any non-numeric characters except dot and minus
    s = re.sub(r"[^0-9.\-]", "", s)
    if not s:
        return default
    try:
        return float(s)
    except (ValueError, TypeError):
        return default


def col_letter_to_index(letter: str) -> int:
    """Convert column letter (A, B, ..., Z, AA, ...) to 1-based index."""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def index_to_col_letter(index: int) -> str:
    """Convert 1-based column index to letter (1=A, 2=B, ...)."""
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_cell(ws, col_letter: str, row: int):
    """Get cell value from worksheet by column letter and row number."""
    return ws.cell(row=row, column=col_letter_to_index(col_letter)).value


def load_mapping(mapping_path: str | Path) -> dict:
    """Load field mapping YAML config."""
    with open(mapping_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def _should_skip_column(header: str, skip_keywords: list[str]) -> bool:
    """Check if a column header matches any skip keywords (metadata columns)."""
    if not header:
        return True
    header_lower = header.lower()
    for keyword in skip_keywords:
        if keyword.lower() in header_lower:
            return True
    return False


def _strip_prefix(text: str, prefix: str) -> str:
    """Strip segment prefix (> or >>) from sub-segment text."""
    stripped = text.strip()
    while stripped.startswith(prefix):
        stripped = stripped[len(prefix):]
    return stripped.strip()


def _parse_segments(ws, config: dict) -> list[SegmentData]:
    """Dynamically discover and parse all segment columns."""
    start_col_idx = col_letter_to_index(config["start_col"])
    header_row = config["header_row"]
    data_start_row = config["data_start_row"]
    prefix = config.get("sub_segment_prefix", ">")
    skip_keywords = config.get("skip_col_keywords", ["%", "Market share"])

    segments = []
    col_idx = start_col_idx

    # Scan rightward across header row until we find an empty column
    # (with a tolerance of 2 empty columns to handle gaps)
    empty_count = 0
    while empty_count < 3:
        header_val = ws.cell(row=header_row, column=col_idx).value
        if not header_val:
            empty_count += 1
            col_idx += 1
            continue
        empty_count = 0

        header_str = str(header_val).strip()

        # Skip metadata columns (% Market share, etc.)
        if _should_skip_column(header_str, skip_keywords):
            col_idx += 1
            continue

        # This is a segment column — collect sub-segments downward
        sub_segments = []
        row = data_start_row
        while True:
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val is None:
                break
            text = _strip_prefix(str(cell_val), prefix)
            if text:
                sub_segments.append(text)
            row += 1

        # Dominating segment = first/top sub-segment
        dominating = sub_segments[0] if sub_segments else ""

        segments.append(SegmentData(
            name=header_str,
            sub_segments=sub_segments,
            dominating=dominating,
        ))

        col_idx += 1

    return segments


def _parse_labeled_rows(ws, config: dict) -> list[str]:
    """Scan column B for rows matching a label, return descriptions from description_col."""
    label = config["search_label"].lower()
    desc_col_idx = col_letter_to_index(config["description_col"])
    start_row = config["start_row"]
    end_row = config["end_row"]

    results = []
    for row in range(start_row, end_row + 1):
        cell_b = ws.cell(row=row, column=2).value  # Column B
        if cell_b and label in str(cell_b).lower():
            desc = ws.cell(row=row, column=desc_col_idx).value
            if desc:
                results.append(str(desc).strip())
    return results


def _parse_regions(ws, config: dict) -> tuple[str, str]:
    """Find dominating and fastest growing regions."""
    start_row = config["start_row"]
    end_row = config["end_row"]
    name_col_idx = col_letter_to_index(config["region_name_col"])
    status_col_idx = col_letter_to_index(config["status_col"])

    dominating = ""
    fastest_growing = ""

    for row in range(start_row, end_row + 1):
        status = ws.cell(row=row, column=status_col_idx).value
        name = ws.cell(row=row, column=name_col_idx).value
        if not status or not name:
            continue
        status_str = str(status).strip().lower()
        name_str = str(name).strip()
        if "dominating" in status_str:
            dominating = name_str
        elif "fastest" in status_str:
            fastest_growing = name_str

    return dominating, fastest_growing


def _parse_market_size(ws, fields: dict, template: str) -> str:
    """Build the market size text from Excel data using the template."""
    row1 = fields["market_size_row1"]
    row2 = fields["market_size_row2"]
    cagr_row = fields["cagr_row"]

    # Read labels from column B to extract years
    label1 = str(ws.cell(row=row1, column=2).value or "")
    label2 = str(ws.cell(row=row2, column=2).value or "")

    # Extract years from labels like "Market size in 2026"
    year1_match = re.search(r"\d{4}", label1)
    year2_match = re.search(r"\d{4}", label2)
    year1 = year1_match.group() if year1_match else "N/A"
    year2 = year2_match.group() if year2_match else "N/A"

    # Read values from column D
    size1 = str(ws.cell(row=row1, column=4).value or "N/A")
    size2 = str(ws.cell(row=row2, column=4).value or "N/A")

    # Read CAGR from column D
    cagr_raw = ws.cell(row=cagr_row, column=4).value
    if isinstance(cagr_raw, (int, float)):
        cagr = f"{cagr_raw * 100:.1f}" if cagr_raw < 1 else f"{cagr_raw:.1f}"
    else:
        cagr = str(cagr_raw or "N/A")

    return template.format(
        year1=year1, size1=size1,
        year2=year2, size2=size2,
        cagr=cagr,
    )


def parse_excel(file_path: str | Path, mapping_path: str | Path = None) -> MarketData:
    """
    Parse an Excel file into a MarketData object using the field mapping config.
    Works with any Excel file that follows the standard CMI format.
    """
    if mapping_path is None:
        mapping_path = Path(__file__).parent.parent / "config" / "field_mapping.yaml"

    mapping = load_mapping(mapping_path)
    sheet_name = mapping["sheet"]
    fields = mapping["fields"]

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Parse market name
    market_name_cell = fields["market_name"]
    col_letter = re.match(r"[A-Z]+", market_name_cell).group()
    row_num = int(re.search(r"\d+", market_name_cell).group())
    market_name = str(get_cell(ws, col_letter, row_num) or "")

    # Parse market size text
    template = mapping.get("market_size_template",
                           "Market size in {year1} is {size1} and for {year2} is {size2} and CAGR is {cagr}%")
    market_size_text = _parse_market_size(ws, fields, template)

    # Disease name
    disease_name = str(fields.get("disease_name", "NA"))

    # Parse segments dynamically
    segments = _parse_segments(ws, mapping["segments"])

    # Parse drivers
    drivers = _parse_labeled_rows(ws, mapping["drivers"])
    driver_1 = drivers[0] if len(drivers) > 0 else ""
    driver_2 = drivers[1] if len(drivers) > 1 else ""

    # Parse restraints
    restraints = _parse_labeled_rows(ws, mapping["restraints"])
    restrain = restraints[0] if restraints else ""

    # Parse opportunities
    opportunities = _parse_labeled_rows(ws, mapping["opportunities"])
    opportunity = opportunities[0] if opportunities else ""

    # Parse regions
    dominating_region, fastest_growing_region = _parse_regions(ws, mapping["regions"])

    wb.close()

    return MarketData(
        market_name=market_name,
        market_size_text=market_size_text,
        disease_name=disease_name,
        segments=segments,
        driver_1=driver_1,
        driver_2=driver_2,
        restrain=restrain,
        opportunity=opportunity,
        dominating_region=dominating_region,
        fastest_growing_region=fastest_growing_region,
    )


def _parse_chart_data(ws, config: dict) -> list[ChartItem]:
    """Extract parent-level segment data for doughnut chart.

    Filters H:I columns to only include items with exactly 1 '>' prefix
    (parent segments, not sub-segments with '>>' or deeper).
    """
    start_col_idx = col_letter_to_index(config["start_col"])
    data_start_row = config["data_start_row"]
    prefix = config.get("sub_segment_prefix", ">")

    # Find the first actual segment column (skip metadata columns)
    skip_keywords = config.get("skip_col_keywords", ["%", "Market share"])
    col_idx = start_col_idx
    value_col_idx = None

    header_val = ws.cell(row=config["header_row"], column=col_idx).value
    if header_val and not _should_skip_column(str(header_val), skip_keywords):
        # Next column should be the market share % column
        value_col_idx = col_idx + 1

    if value_col_idx is None:
        return []

    items = []
    row = data_start_row
    while True:
        h_val = ws.cell(row=row, column=col_idx).value
        i_val = ws.cell(row=row, column=value_col_idx).value
        if h_val is None:
            break
        h_str = str(h_val).strip()
        indent_level = len(h_str) - len(h_str.lstrip(prefix))
        if indent_level == 1:  # Parent-level segments only (single >)
            label = h_str.lstrip(prefix).strip()
            try:
                value = float(i_val) if i_val is not None else 0.0
            except (ValueError, TypeError):
                value = 0.0
            items.append(ChartItem(label=label, value=value))
        row += 1

    return items


def _parse_geographic_data(wb, sheet_name: str = "Geographies") -> dict[str, list[str]]:
    """Parse the Geographies sheet into {region: [countries...]} dict."""
    if sheet_name not in wb.sheetnames:
        return {}

    ws = wb[sheet_name]
    geo_data = {}

    # Row 1 = region headers, rows 2+ = countries
    for col in range(1, 20):
        region = ws.cell(row=1, column=col).value
        if not region:
            break
        region = str(region).strip()
        countries = []
        for row in range(2, 20):
            country = ws.cell(row=row, column=col).value
            if not country:
                break
            countries.append(str(country).strip())
        geo_data[region] = countries

    return geo_data


def parse_excel_full(
    file_path: str | Path,
    mapping_path: str | Path = None,
) -> ExtendedMarketData:
    """Parse Excel file into ExtendedMarketData with all fields for PPTX/DOCX generation."""
    if mapping_path is None:
        mapping_path = Path(__file__).parent.parent / "config" / "field_mapping.yaml"

    mapping = load_mapping(mapping_path)
    sheet_name = mapping["sheet"]

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    ws = wb[sheet_name]

    # --- Basic fields (same as parse_excel) ---
    market_name = str(get_cell(ws, "D", 2) or "")
    template = mapping.get("market_size_template",
                           "Market size in {year1} is {size1} and for {year2} is {size2} and CAGR is {cagr}%")
    market_size_text = _parse_market_size(ws, mapping["fields"], template)
    disease_name = str(mapping["fields"].get("disease_name", "NA"))
    segments = _parse_segments(ws, mapping["segments"])

    # Drivers
    drivers = _parse_labeled_rows(ws, mapping["drivers"])
    driver_1 = drivers[0] if len(drivers) > 0 else ""
    driver_2 = drivers[1] if len(drivers) > 1 else ""

    # Restraints
    restraints = _parse_labeled_rows(ws, mapping["restraints"])
    restrain = restraints[0] if restraints else ""

    # Opportunities
    opportunities = _parse_labeled_rows(ws, mapping["opportunities"])
    opportunity = opportunities[0] if opportunities else ""

    # Regions (dominating/fastest)
    dominating_region, fastest_growing_region = _parse_regions(ws, mapping["regions"])

    # --- Extended fields ---
    # Impact analysis indicators
    driver_1_indicator = int(get_cell(ws, "D", 15) or 0)
    driver_2_indicator = int(get_cell(ws, "D", 16) or 0)
    restraint_1 = str(get_cell(ws, "C", 17) or "")
    restraint_2 = str(get_cell(ws, "C", 18) or "")
    restraint_1_indicator = int(get_cell(ws, "D", 17) or 0)
    restraint_2_indicator = int(get_cell(ws, "D", 18) or 0)
    opportunity_1 = str(get_cell(ws, "C", 19) or "")
    opportunity_2 = str(get_cell(ws, "C", 20) or "")
    opportunity_1_indicator = int(get_cell(ws, "D", 19) or 0)
    opportunity_2_indicator = int(get_cell(ws, "D", 20) or 0)

    # Market metrics
    market_share_raw = get_cell(ws, "D", 22)
    market_share_pct = _safe_float(market_share_raw)
    market_size_value = str(get_cell(ws, "D", 7) or "")
    forecast_size_value = str(get_cell(ws, "D", 8) or "")
    base_year = int(get_cell(ws, "D", 4) or 2025)
    historical_start = int(get_cell(ws, "D", 5) or 2020)
    forecast_end = int(get_cell(ws, "D", 6) or 2033)
    cagr_raw = get_cell(ws, "D", 9)
    cagr_value = _safe_float(cagr_raw)
    currency_type = str(get_cell(ws, "D", 10) or "USD")

    # Concentration
    concentration_value = int(get_cell(ws, "D", 31) or 0)

    # Companies from column G
    companies = []
    for row in range(3, 50):
        val = ws.cell(row=row, column=7).value
        if not val:
            break
        companies.append(str(val).strip())

    # Takeaways from D40:D44
    takeaways = []
    for row in range(40, 45):
        val = ws.cell(row=row, column=4).value
        if val:
            takeaways.append(str(val).strip())

    # Regions with status
    regions = []
    for row in range(23, 29):
        name = ws.cell(row=row, column=3).value
        status = ws.cell(row=row, column=4).value
        if name:
            regions.append(RegionData(
                name=str(name).strip(),
                status=str(status).strip() if status and str(status).strip() != "-" else "",
            ))

    # Chart data (filtered parent-level segments)
    segment_chart_data = _parse_chart_data(ws, mapping["segments"])

    # Geographic data from Geographies sheet
    geographic_data = _parse_geographic_data(wb)

    wb.close()

    return ExtendedMarketData(
        market_name=market_name,
        market_size_text=market_size_text,
        disease_name=disease_name,
        segments=segments,
        driver_1=driver_1,
        driver_2=driver_2,
        restrain=restrain,
        opportunity=opportunity,
        dominating_region=dominating_region,
        fastest_growing_region=fastest_growing_region,
        driver_1_indicator=driver_1_indicator,
        driver_2_indicator=driver_2_indicator,
        restraint_1=restraint_1,
        restraint_2=restraint_2,
        restraint_1_indicator=restraint_1_indicator,
        restraint_2_indicator=restraint_2_indicator,
        opportunity_1=opportunity_1,
        opportunity_2=opportunity_2,
        opportunity_1_indicator=opportunity_1_indicator,
        opportunity_2_indicator=opportunity_2_indicator,
        market_share_pct=market_share_pct,
        market_size_value=market_size_value,
        forecast_size_value=forecast_size_value,
        base_year=base_year,
        historical_start=historical_start,
        forecast_end=forecast_end,
        cagr_value=cagr_value,
        currency_type=currency_type,
        concentration_value=concentration_value,
        companies=companies,
        takeaways=takeaways,
        regions=regions,
        segment_chart_data=segment_chart_data,
        geographic_data=geographic_data,
    )
